"""Resource-bounded extraction for untrusted XLSX and text PDF supplier quotes."""

from __future__ import annotations

import io
import re
import time
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

PARSER_VERSION = "packaging-quote-v3"
MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_XLSX_ENTRIES = 2_000
MAX_XLSX_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
MAX_XLSX_SHEETS = 5
MAX_XLSX_ROWS = 500
MAX_XLSX_COLUMNS = 40
MAX_PDF_PAGES = 20
MAX_EXTRACTED_CHARS = 200_000
MAX_TEXT_VALUE_LENGTH = 2000
REVIEW_THRESHOLD = 0.80


class QuoteParseError(ValueError):
    """A quote cannot be parsed within the supported trust and resource boundary."""


FIELD_META: dict[str, dict[str, Any]] = {
    "supplier_name": {"label": "供应商", "kind": "text", "required": True},
    "item_description": {"label": "品名/描述", "kind": "text", "required": True},
    "material": {"label": "材质", "kind": "text", "required": True},
    "color": {"label": "颜色", "kind": "text", "required": True},
    "print_colors": {"label": "印刷色数", "kind": "integer", "required": True},
    "currency": {"label": "币种", "kind": "currency", "required": True},
    "unit_price": {"label": "报价", "kind": "decimal", "required": True},
    "price_basis": {"label": "计价数量", "kind": "integer", "required": True},
    "tax_rate": {"label": "税率", "kind": "rate", "required": True},
    "tax_included": {"label": "是否含税", "kind": "boolean", "required": True},
    "shipping_fee": {"label": "运费", "kind": "decimal", "required": False},
    "shipping_included": {"label": "是否含运费", "kind": "boolean", "required": True},
    "moq": {"label": "起订量（MOQ）", "kind": "integer", "required": True},
    "lead_time_days": {"label": "交期（天）", "kind": "integer", "required": True},
    "supports_invoice": {"label": "是否可开票", "kind": "boolean", "required": True},
    "width_mm": {"label": "宽度（mm）", "kind": "decimal", "required": True},
    "length_mm": {"label": "长度（mm）", "kind": "decimal", "required": True},
    "thickness_um": {"label": "厚度（µm）", "kind": "decimal", "required": True},
    "payment_terms": {"label": "付款条件", "kind": "text", "required": False},
    "valid_until": {"label": "报价有效期", "kind": "date", "required": False},
}

_ALIASES = {
    "supplier_name": ["供应商", "供应商名称", "报价方", "公司名称", "vendor", "supplier", "suppliername"],
    "item_description": ["品名", "产品名称", "物料名称", "物料描述", "产品描述", "规格描述", "description", "item", "product"],
    "material": ["材质", "材料", "原料", "material"],
    "color": ["颜色", "色彩", "color", "colour"],
    "print_colors": ["印刷色数", "印刷颜色数", "印刷", "printcolors", "printingcolors"],
    "currency": ["币种", "货币", "currency"],
    "unit_price": ["单价", "报价", "含税单价", "未税单价", "price", "unitprice", "quotedprice"],
    "price_basis": ["计价数量", "计价单位", "报价单位", "价格单位", "pricebasis", "priceunit", "per"],
    "tax_rate": ["税率", "增值税率", "vat", "taxrate"],
    "tax_included": ["是否含税", "含税", "税费包含", "taxincluded", "includestax"],
    "shipping_fee": ["运费", "物流费", "配送费", "freight", "shipping", "shippingfee"],
    "shipping_included": ["是否含运费", "运费包含", "是否包邮", "shippingincluded", "freightincluded"],
    "moq": ["moq", "起订量", "最小起订量", "minimumorder", "minimumorderquantity"],
    "lead_time_days": ["交期", "交货期", "生产周期", "交期天", "leaddays", "leadtime", "deliverydays"],
    "supports_invoice": [
        "可开票",
        "是否可开票",
        "可开专票",
        "是否可开专票",
        "可开普票",
        "是否可开普票",
        "专票",
        "普票",
        "增值税专票",
        "增值税专用发票",
        "增值税普通发票",
        "发票",
        "invoice",
        "supportsinvoice",
    ],
    "width_mm": ["宽", "宽度", "宽mm", "宽度mm", "width", "widthmm"],
    "length_mm": ["长", "长度", "高度", "长mm", "长度mm", "length", "height", "lengthmm"],
    "thickness_um": ["厚度", "厚度um", "厚度微米", "丝数", "thickness", "thicknessum", "micron"],
    "payment_terms": ["付款条件", "结算方式", "账期", "payment", "paymentterms"],
    "valid_until": ["有效期", "报价有效期", "validuntil", "expiry", "expiration"],
}


def _key(value: Any) -> str:
    text = str(value or "").strip().lower().replace("μ", "u").replace("µ", "u")
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


ALIASES = {_key(alias): field for field, aliases in _ALIASES.items() for alias in aliases}


def _plain_decimal(value: Decimal) -> str:
    normalized = value.normalize()
    return format(normalized, "f") if normalized != 0 else "0"


def _decimal(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            result = Decimal(str(value))
            return _plain_decimal(result) if result.is_finite() else None
        except InvalidOperation:
            return None
    text = str(value).strip().replace(",", "")
    match = re.search(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        result = Decimal(match.group(0))
        return _plain_decimal(result) if result.is_finite() else None
    except InvalidOperation:
        return None


def _integer(value: Any) -> int | None:
    number = _decimal(value)
    if number is None:
        return None
    try:
        result = Decimal(number)
        if result != result.to_integral_value():
            return None
        return int(result)
    except (InvalidOperation, ValueError, OverflowError):
        return None


_COLOR_CANONICAL = {
    "black": "black",
    "黑色": "black",
    "white": "white",
    "白色": "white",
    "transparent": "transparent",
    "clear": "transparent",
    "透明": "transparent",
    "red": "red",
    "红色": "red",
    "blue": "blue",
    "蓝色": "blue",
}


def _values_equivalent(field: str, current: Any, incoming: Any) -> bool:
    if current == incoming:
        return True
    if field == "color":
        current_key = _COLOR_CANONICAL.get(_key(current), _key(current))
        incoming_key = _COLOR_CANONICAL.get(_key(incoming), _key(incoming))
        return bool(current_key) and current_key == incoming_key
    return False


def _boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    text = _key(value)
    # 运费措辞歧义：同时出现“包邮”与“到付/自付/另算/自理”等负向标记时
    # （如“江浙沪包邮，新疆西藏运费到付”）不能直接判为含运费，必须人工复核。
    # 先剔除负向短语，避免“不包邮”中的“包邮”被误当成正向标记。
    remaining = text
    for token in _SHIPPING_NEGATIVE_TOKENS:
        remaining = remaining.replace(token, "")
    if any(token in text for token in _SHIPPING_NEGATIVE_TOKENS) and any(
        token in remaining for token in _SHIPPING_POSITIVE_TOKENS
    ):
        return None
    if any(
        token in text
        for token in (
            "不包邮",
            "到付",
            "自付",
            "另算",
            "自理",
            "不含运费",
            "运费另计",
            "不可开票",
            "不能开票",
            "不开票",
            "不提供发票",
            "不提供专票",
            "不支持开票",
            "不含",
            "另计",
            "excluded",
            "notincluded",
            "noinvoice",
            "invoiceunavailable",
        )
    ):
        return False
    if text in {"是", "有", "支持", "含", "包含", "包邮", "yes", "y", "true", "included", "include", "1"}:
        return True
    if text in {"否", "无", "不支持", "不含", "另计", "no", "n", "false", "excluded", "exclude", "0"}:
        return False
    if any(token in text for token in ("含税", "包邮", "included")):
        return True
    return None


_SHIPPING_POSITIVE_TOKENS = (
    "包邮",
    "运费已含",
    "含运费",
    "运费包含",
    "shippingincluded",
    "freightincluded",
)
_SHIPPING_NEGATIVE_TOKENS = (
    "不包邮",
    "到付",
    "自付",
    "另算",
    "自理",
    "不含运费",
    "运费另计",
    "运费不含",
    "运费未含",
    "shippingnotincluded",
    "freightnotincluded",
    "shippingexcluded",
    "freightexcluded",
)


_INVOICE_NEGATIVE = (
    "不可开票",
    "不能开票",
    "不开票",
    "无法开票",
    "不可开专票",
    "不能开专票",
    "不开专票",
    "无法开专票",
    "不可开普票",
    "不能开普票",
    "不开普票",
    "无法开普票",
    "不可开具专票",
    "不能开具专票",
    "无法开具专票",
    "不可开具普票",
    "不能开具普票",
    "无法开具普票",
    "不支持开票",
    "不支持开专票",
    "不支持开普票",
    "不支持开具专票",
    "不支持开具普票",
    "不提供发票",
    "不提供专票",
    "不提供普票",
    "不提供增值税专用发票",
    "不提供增值税普通发票",
    "不能开具增值税专用发票",
    "不可开具增值税专用发票",
    "无法开具增值税专用发票",
    "不能开具增值税普通发票",
    "不可开具增值税普通发票",
    "无法开具增值税普通发票",
    "不能开增值税专用发票",
    "不开增值税专用发票",
    "无法开增值税专用发票",
    "不能开增值税普通发票",
    "不开增值税普通发票",
    "无法开增值税普通发票",
    "noinvoice",
    "invoiceunavailable",
)
_INVOICE_POSITIVE = (
    "可开",
    "专票",
    "普票",
    "能开",
    "可开发票",
    "invoiceavailable",
    "invoicesupport",
)


def _supports_invoice(value: Any) -> bool | None:
    """Boolean parse for supports_invoice that prioritises invoice-specific
    markers over generic negation tokens such as 不含税 ("不含税可开专票" must
    be treated as invoice-capable, not as "cannot invoice")."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    text = _key(value)
    if any(token in text for token in _INVOICE_NEGATIVE):
        return False
    if any(token in text for token in _INVOICE_POSITIVE):
        return True
    return _boolean(text)


def _currency(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if not text:
        return None
    if "USD" in text or "$" in text or "美元" in text:
        return "USD"
    if "EUR" in text or "€" in text or "欧元" in text:
        return "EUR"
    if "CNY" in text or "RMB" in text or "¥" in text or "￥" in text or "人民币" in text:
        return "CNY"
    return text if re.fullmatch(r"[A-Z]{3}", text) else None


def _rate(value: Any) -> str | None:
    number = _decimal(value)
    if number is None:
        return None
    result = Decimal(number)
    if "%" in str(value) or result > 1:
        result /= Decimal("100")
    if result < 0 or result > 1:
        return None
    return _plain_decimal(result)


def _basis(value: Any) -> int | None:
    text = str(value or "").lower().replace(",", "")
    if (
        "万" in text
        or re.search(r"(?<!\d)10000(?!\d)", text)
        or re.search(r"(?<!\d)10k(?![a-z\d])", text)
        or "ten thousand" in text
    ):
        return 10_000
    if (
        "千" in text
        or re.search(r"(?<!\d)1000(?!\d)", text)
        or re.search(r"(?<!\d)1k(?![a-z\d])", text)
        or "thousand" in text
    ):
        return 1000
    if "百" in text or re.search(r"(?<!\d)100(?!\d)", text) or "hundred" in text:
        return 100
    if any(token in text for token in ("单个", "每个", "/个", "piece", "each", "per pc")):
        return 1
    return _integer(value)


def _date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    for pattern in (r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", r"(20\d{2})(\d{2})(\d{2})"):
        match = re.search(pattern, text)
        if match:
            try:
                return date(*(int(part) for part in match.groups())).isoformat()
            except ValueError:
                return None
    return None


def coerce_field_value(field: str, value: Any) -> Any:
    if field not in FIELD_META:
        raise ValueError(f"不支持的报价字段：{field}")
    kind = FIELD_META[field]["kind"]
    if kind == "decimal":
        result = _decimal(value)
        if result is None:
            return None
        number = Decimal(result)
        if field in {"unit_price", "width_mm", "length_mm", "thickness_um"} and number <= 0:
            return None
        if field == "shipping_fee" and number < 0:
            return None
        return result
    if kind == "integer":
        result = _basis(value) if field == "price_basis" else _integer(value)
        if result is None:
            return None
        if field in {"price_basis", "moq"} and result <= 0:
            return None
        if field == "lead_time_days" and result < 0:
            return None
        if field == "print_colors" and not 0 <= result <= 12:
            return None
        return result
    if kind == "boolean":
        return (
            _supports_invoice(value)
            if field == "supports_invoice"
            else _boolean(value)
        )
    if kind == "currency":
        return _currency(value)
    if kind == "rate":
        return _rate(value)
    if kind == "date":
        return _date(value)
    text = str(value or "").strip()
    if len(text) > MAX_TEXT_VALUE_LENGTH:
        raise ValueError(
            f"{FIELD_META[field]['label']} 长度不能超过 {MAX_TEXT_VALUE_LENGTH} 个字符"
        )
    return text or None


def _entry(
    field: str,
    value: Any,
    confidence: float,
    *,
    document_kind: str,
    locator: str,
    excerpt: str,
    method: str,
) -> dict[str, Any]:
    try:
        coerced = coerce_field_value(field, value)
    except ValueError:
        # 超长文本等不可入库的值按无法复核处理（needs_review），避免解析
        # 阶段把 ValueError 泄漏成 500。
        coerced = None
    bounded_confidence = round(max(0.0, min(1.0, confidence)), 2)
    return {
        "value": coerced,
        "confidence": bounded_confidence,
        "status": "accepted" if coerced is not None and bounded_confidence >= REVIEW_THRESHOLD else "needs_review",
        "source": {
            "document_kind": document_kind,
            "locator": locator,
            "excerpt": str(excerpt).strip()[:500],
            "method": method,
        },
    }


def _set_field(fields: dict[str, Any], field: str, entry: dict[str, Any]) -> None:
    current = fields.get(field)
    if current is None:
        fields[field] = entry
        return

    current_confidence = float(current.get("confidence", 0))
    incoming_confidence = float(entry.get("confidence", 0))
    current_value = current.get("value")
    incoming_value = entry.get("value")
    if (
        current_value is not None
        and incoming_value is not None
        and not _values_equivalent(field, current_value, incoming_value)
        and current_confidence >= REVIEW_THRESHOLD
        and incoming_confidence >= REVIEW_THRESHOLD
    ):
        preferred = entry if incoming_confidence > current_confidence else current
        candidates = list(current.get("conflicts") or [])
        candidates.extend(
            [
                {
                    "value": current_value,
                    "confidence": current_confidence,
                    "source": current.get("source", {}),
                },
                {
                    "value": incoming_value,
                    "confidence": incoming_confidence,
                    "source": entry.get("source", {}),
                },
            ]
        )
        unique: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for candidate in candidates:
            source = candidate.get("source") or {}
            key = (repr(candidate.get("value")), str(source.get("locator") or ""))
            if key not in seen:
                seen.add(key)
                unique.append(candidate)
        fields[field] = {
            **preferred,
            "status": "needs_review",
            "conflicts": unique,
        }
        return

    if incoming_confidence > current_confidence:
        fields[field] = entry



def _informational_entry(
    value: Any,
    confidence: float,
    *,
    document_kind: str,
    locator: str,
    excerpt: str,
    method: str,
) -> dict[str, Any]:
    """Evidence entry for a label/value pair that is not part of the fixed
    field set. Stored read-only under ``informational_fields`` so unknown
    fields from rich supplier quotes are never silently dropped."""
    return {
        "value": str(value).strip(),
        "confidence": round(max(0.0, min(1.0, confidence)), 2),
        "status": "accepted",
        "informational": True,
        "label": str(excerpt).split(":", 1)[0].split("?", 1)[0].strip()[:100],
        "source": {
            "document_kind": document_kind,
            "locator": locator,
            "excerpt": str(excerpt).strip()[:500],
            "method": method,
        },
    }


def _informational_key(label: Any) -> str | None:
    key = _key(label)
    if not key or key[0].isdigit():
        return None
    return key


def _set_informational(
    informational: dict[str, Any],
    label: Any,
    entry: dict[str, Any],
) -> None:
    key = _informational_key(label)
    if key is None:
        return
    current = informational.get(key)
    if current is None:
        informational[key] = entry
        return
    if float(entry.get("confidence", 0)) > float(current.get("confidence", 0)):
        informational[key] = entry


def _extract_specs(fields: dict[str, Any], document_kind: str) -> None:
    description = fields.get("item_description")
    if not description or not description.get("value"):
        return
    text = str(description["value"])
    size = re.search(
        r"(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)\s*(mm|cm)?",
        text,
        re.I,
    )
    thickness = re.search(
        r"(\d+(?:\.\d+)?)\s*(um|[μµ]m|微米|micron|丝)", text, re.I
    )
    source = description.get("source", {})
    if size:
        # 单位换算：报价描述里常见的 cm 需转成 mm（1 cm = 10 mm），
        # 避免把 20*30cm 当作 20×30mm 直接接受。
        unit = (size.group(3) or "").lower()
        scale = Decimal("10") if unit == "cm" else Decimal("1")
        for field, raw in zip(("width_mm", "length_mm"), size.groups()[:2], strict=True):
            value = str(Decimal(str(raw)) * scale)
            _set_field(
                fields,
                field,
                _entry(
                    field,
                    value,
                    0.84,
                    document_kind=document_kind,
                    locator=str(source.get("locator", "description")),
                    excerpt=text,
                    method="spec_pattern",
                ),
            )
    if thickness:
        # 中文包装报价常用“丝”：1 丝 = 10 µm，必须换算而不是按 µm 直接接受。
        raw, unit = thickness.groups()
        value = str(Decimal(str(raw)) * (Decimal("10") if unit == "丝" else Decimal("1")))
        _set_field(
            fields,
            "thickness_um",
            _entry(
                "thickness_um",
                value,
                0.84,
                document_kind=document_kind,
                locator=str(source.get("locator", "description")),
                excerpt=text,
                method="spec_pattern",
            ),
        )

    material_match = re.search(
        r"(?<![A-Za-z])(PE|PVC|PP|PET|PLA)(?![A-Za-z])|聚乙烯|聚氯乙烯|聚丙烯|聚乳酸",
        text,
        re.I,
    )
    if material_match:
        raw_material = material_match.group(0)
        material_aliases = {
            "聚乙烯": "PE",
            "聚氯乙烯": "PVC",
            "聚丙烯": "PP",
            "聚乳酸": "PLA",
        }
        material = material_aliases.get(raw_material, raw_material.upper())
        _set_field(
            fields,
            "material",
            _entry(
                "material",
                material,
                0.86,
                document_kind=document_kind,
                locator=str(source.get("locator", "description")),
                excerpt=text,
                method="identity_pattern",
            ),
        )

    color_patterns = (
        (r"白色|(?<![A-Za-z])white(?![A-Za-z])", "白色"),
        (r"黑色|(?<![A-Za-z])black(?![A-Za-z])", "黑色"),
        (r"透明|(?<![A-Za-z])(?:transparent|clear)(?![A-Za-z])", "透明"),
        (r"红色|(?<![A-Za-z])red(?![A-Za-z])", "红色"),
        (r"蓝色|(?<![A-Za-z])blue(?![A-Za-z])", "蓝色"),
    )
    color = next((value for pattern, value in color_patterns if re.search(pattern, text, re.I)), None)
    if color:
        _set_field(
            fields,
            "color",
            _entry(
                "color",
                color,
                0.86,
                document_kind=document_kind,
                locator=str(source.get("locator", "description")),
                excerpt=text,
                method="identity_pattern",
            ),
        )

    print_colors: int | None = None
    if re.search(r"无印刷|不印刷|素袋|unprinted|no\s+print", text, re.I):
        print_colors = 0
    elif re.search(r"单色|一色|one[-\s]?colou?r", text, re.I):
        print_colors = 1
    elif re.search(r"双色|二色|two[-\s]?colou?r", text, re.I):
        print_colors = 2
    else:
        print_match = re.search(r"(\d{1,2})\s*(?:色|colou?rs?)", text, re.I)
        if print_match:
            print_colors = int(print_match.group(1))
    if print_colors is not None:
        _set_field(
            fields,
            "print_colors",
            _entry(
                "print_colors",
                print_colors,
                0.86,
                document_kind=document_kind,
                locator=str(source.get("locator", "description")),
                excerpt=text,
                method="identity_pattern",
            ),
        )


def _infer_common(fields: dict[str, Any], text: str, document_kind: str) -> None:
    def inferred(field: str, value: Any, confidence: float, excerpt: str) -> None:
        _set_field(
            fields,
            field,
            _entry(
                field,
                value,
                confidence,
                document_kind=document_kind,
                locator="document text",
                excerpt=excerpt,
                method="document_pattern",
            ),
        )

    shipping_negative = re.search(
        r"不包邮|(?:运费)?(?:不含(?!税)|未含(?!税)|另计|另算|到付|自付|自理)|不含运费|"
        r"shipping\s+(?:is\s+)?(?:not\s+included|excluded)|"
        r"freight\s+(?:is\s+)?(?:not\s+included|excluded)",
        text,
        re.I,
    )
    shipping_positive = re.search(
        r"(?<!不)包邮|运费已含|报价含运费|价格含运费|已含运费|"
        r"shipping\s+is\s+included|freight\s+is\s+included",
        text,
        re.I,
    )
    if shipping_negative and shipping_positive:
        # 同时出现“包邮”与“到付/自付/另算/自理”等负向标记（如
        # “江浙沪包邮，新疆西藏运费到付”）：语义冲突，不能静默判为含运费
        # 并置运费为 0，必须人工复核。
        excerpt = (
            f"{shipping_positive.group(0)}；{shipping_negative.group(0)}"
        )
        inferred("shipping_included", None, 0.3, excerpt)
    elif shipping_negative:
        # “运费不含税”是关于税的口径，不是“运费另计”，(?!税) 排除该误命中；
        # 摘录使用原文命中的片段而不是硬编码文案，保证审计证据与原件一致。
        inferred(
            "shipping_included", False, 0.91, shipping_negative.group(0)
        )
    elif shipping_positive:
        inferred("shipping_included", True, 0.91, shipping_positive.group(0))
        inferred("shipping_fee", "0", 0.91, shipping_positive.group(0))
    if re.search(
        r"价格含税|报价含税|已含税|tax\s+is\s+included|vat\s+is\s+included",
        text,
        re.I,
    ):
        inferred("tax_included", True, 0.88, "含税")
    if re.search(r"未税|不含税|tax\s+excluded|vat\s+excluded", text, re.I):
        inferred("tax_included", False, 0.91, "不含税")
    current_invoice = fields.get("supports_invoice")
    invoice_confirmed = (
        isinstance(current_invoice, dict)
        and current_invoice.get("status") == "accepted"
        and float(current_invoice.get("confidence", 0)) >= REVIEW_THRESHOLD
    )
    if re.search(
        r"不可开票|不能开票|不开票|无法开票|不可开专票|不能开专票|不开专票|无法开专票|"
        r"不可开普票|不能开普票|不开普票|无法开普票|"
        r"(?:不可|不能|无法|不开|不支持|不提供)(?:开具|开)?增值税(?:专用|普通)?发票|"
        r"(?:不可|不能|无法|不开|不支持|不提供)(?:开具)?(?:专票|普票)|"
        r"不(?:提供|支持)(?:发票|专票|普票|开票)|"
        r"no[ \t]+invoice|invoice[ \t]*:?[ \t]*(?:no|unavailable)",
        text,
        re.I,
    ):
        inferred("supports_invoice", False, 0.91, "不可开票")
    elif not invoice_confirmed and re.search(
        r"(?<!是否)可开|专票|普票|invoice\s*:?\s*(yes|available)",
        text,
        re.I,
    ):
        # 正向推断只在高置信度显式结果不存在时进行：
        # - “是否可开票”标签自身的“可开”由负向后行排除；
        # - “可开票/专票/普票”作为表头或标签时（如“可开票: 否”），显式值已
        #   高置信度接受，不再用标签文本制造虚假的跨来源冲突；
        # 反向证据（“不可开票/不开专票”等）始终保留，用于发现真实矛盾
        # （例如显式“是”+正文“本公司不可开票”）。
        inferred("supports_invoice", True, 0.88, "可开票")
    if "currency" not in fields:
        currency = _currency(text)
        if currency:
            inferred("currency", currency, 0.82, currency)
    _extract_specs(fields, document_kind)


def _finalize(
    fields: dict[str, Any],
    *,
    filename: str,
    document_kind: str,
    processing_ms: float,
    informational: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if "supplier_name" not in fields:
        stem = Path(filename).stem
        # 演示/真实报价文件名常带“报价单/QUOTATION”等通用后缀；这些不是供应商名
        # 的一部分，去掉后 fallback 才与真值一致（例如 星河包装报价单 -> 星河包装）。
        stem = re.sub(r"(报价单|报价|QUOTATION|Quotation|quotation)$", "", stem)
        fallback = re.sub(r"[_-]+", " ", stem).strip()
        fields["supplier_name"] = _entry(
            "supplier_name",
            fallback,
            0.55,
            document_kind=document_kind,
            locator="filename",
            excerpt=filename,
            method="filename_fallback",
        )
    for field, meta in FIELD_META.items():
        if meta["required"] and field not in fields:
            fields[field] = _entry(
                field,
                None,
                0.0,
                document_kind=document_kind,
                locator="not found",
                excerpt="",
                method="missing",
            )
    if fields.get("shipping_included", {}).get("value") is False and "shipping_fee" not in fields:
        fields["shipping_fee"] = _entry(
            "shipping_fee",
            None,
            0.0,
            document_kind=document_kind,
            locator="not found",
            excerpt="",
            method="missing",
        )
    return {
        "schema_version": 1,
        "parser_version": PARSER_VERSION,
        "document_kind": document_kind,
        "fields": fields,
        "informational_fields": informational or {},
        "processing_ms": round(processing_ms, 2),
    }


def _validate_file(filename: str, data: bytes) -> str:
    if not filename or Path(filename).name != filename:
        raise QuoteParseError("文件名无效")
    if len(data) == 0:
        raise QuoteParseError("报价文件为空")
    if len(data) > MAX_FILE_BYTES:
        raise QuoteParseError(f"报价文件不得超过 {MAX_FILE_BYTES // 1024 // 1024} MB")
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".pdf"}:
        raise QuoteParseError("仅支持 .xlsx 和文本型 .pdf 报价")
    return suffix


def _validate_xlsx_archive(data: bytes) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise QuoteParseError("XLSX 文件结构无效") from exc
    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ENTRIES:
            raise QuoteParseError("XLSX ZIP 条目过多")
        total = sum(item.file_size for item in entries)
        if total > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise QuoteParseError("XLSX 解压后内容过大")
        for item in entries:
            if item.file_size > 10 * 1024 * 1024:
                raise QuoteParseError("XLSX 内部条目过大")
            if item.compress_size and item.file_size > 1024 * 1024:
                if item.file_size / item.compress_size > 100:
                    raise QuoteParseError("XLSX 压缩比异常")


def _deadline(budget_s: float | None) -> float | None:
    if budget_s is None:
        return None
    return time.monotonic() + budget_s


def _check_deadline(deadline: float | None, *, what: str) -> None:
    if deadline is not None and time.monotonic() >= deadline:
        raise QuoteParseError(f"{what}解析超过时间预算，已中止（请检查文件）")


def _xlsx_quote(
    filename: str, data: bytes, *, time_budget_s: float | None = None
) -> dict[str, Any]:
    started = time.perf_counter()
    _validate_xlsx_archive(data)
    deadline = _deadline(time_budget_s)
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 - normalize parser errors for the API
        raise QuoteParseError("无法读取 XLSX 报价") from exc
    try:
        if len(workbook.worksheets) > MAX_XLSX_SHEETS:
            raise QuoteParseError(f"XLSX 工作表不得超过 {MAX_XLSX_SHEETS} 个")
        fields: dict[str, Any] = {}
        informational: dict[str, Any] = {}
        document_lines: list[str] = []
        for sheet in workbook.worksheets:
            _check_deadline(deadline, what="XLSX")
            if int(sheet.max_row or 0) > MAX_XLSX_ROWS:
                raise QuoteParseError(f"XLSX 每个工作表不得超过 {MAX_XLSX_ROWS} 行")
            if int(sheet.max_column or 0) > MAX_XLSX_COLUMNS:
                raise QuoteParseError(f"XLSX 每个工作表不得超过 {MAX_XLSX_COLUMNS} 列")
            rows = list(
                sheet.iter_rows(
                    min_row=1,
                    max_row=MAX_XLSX_ROWS,
                    min_col=1,
                    max_col=MAX_XLSX_COLUMNS,
                    values_only=True,
                )
            )
            for row_index, row in enumerate(rows, start=1):
                if row_index % 50 == 0:
                    _check_deadline(deadline, what="XLSX")
                populated = [(index + 1, value) for index, value in enumerate(row) if value not in (None, "")]
                if populated:
                    document_lines.append(" | ".join(str(value) for _column, value in populated))
                header_map = {
                    column: ALIASES[_key(value)]
                    for column, value in populated
                    if _key(value) in ALIASES
                }
                if len(header_map) >= 3:
                    data_entry = next(
                        (
                            (candidate_index, candidate)
                            for candidate_index, candidate in enumerate(
                                rows[row_index : min(len(rows), row_index + 5)],
                                start=row_index + 1,
                            )
                            if any(value not in (None, "") for value in candidate)
                        ),
                        None,
                    )
                    if data_entry:
                        data_row_index, data_row = data_entry
                        for column, field in header_map.items():
                            value = data_row[column - 1]
                            if value in (None, ""):
                                continue
                            label = row[column - 1]
                            locator = (
                                f"{sheet.title}!{get_column_letter(column)}{data_row_index}"
                            )
                            _set_field(
                                fields,
                                field,
                                _entry(
                                    field,
                                    value,
                                    0.97,
                                    document_kind="xlsx",
                                    locator=locator,
                                    excerpt=f"{label}: {value}",
                                    method="table_header",
                                ),
                            )
                    continue
                for position, (column, label) in enumerate(populated):
                    candidate = next(
                        (
                            (candidate_column, candidate_value)
                            for candidate_column, candidate_value in populated[position + 1 :]
                            if candidate_column <= column + 3
                        ),
                        None,
                    )
                    if candidate is None:
                        continue
                    value_column, value = candidate
                    locator = f"{sheet.title}!{get_column_letter(value_column)}{row_index}"
                    excerpt = f"{label}: {value}"
                    field = ALIASES.get(_key(label))
                    if field:
                        _set_field(
                            fields,
                            field,
                            _entry(
                                field,
                                value,
                                0.97,
                                document_kind="xlsx",
                                locator=locator,
                                excerpt=excerpt,
                                method="key_value_cell",
                            ),
                        )
                    elif len(populated) == 2:
                        # Only single label/value rows (exactly two populated
                        # cells) count as unknown key-value fields. Wider rows
                        # are table data/headers and would add noise.
                        _set_informational(
                            informational,
                            label,
                            _informational_entry(
                                value,
                                0.9,
                                document_kind="xlsx",
                                locator=locator,
                                excerpt=excerpt,
                                method="key_value_cell",
                            ),
                        )
        full_text = "\n".join(document_lines)[:MAX_EXTRACTED_CHARS]
        _infer_common(fields, full_text, "xlsx")
        return _finalize(
            fields,
            filename=filename,
            document_kind="xlsx",
            processing_ms=(time.perf_counter() - started) * 1000,
            informational=informational,
        )
    finally:
        workbook.close()


def _pdf_delimited_tables(
    fields: dict[str, Any],
    informational: dict[str, Any],
    lines: list[str],
    page_number: int,
) -> None:
    for line_index, line in enumerate(lines):
        headers = [value.strip() for value in line.split("|")]
        header_map = {
            position: ALIASES[_key(label)]
            for position, label in enumerate(headers)
            if _key(label) in ALIASES
        }
        if len(header_map) < 2:
            continue
        data_index = next(
            (
                candidate
                for candidate in range(line_index + 1, min(len(lines), line_index + 4))
                if lines[candidate].strip()
            ),
            None,
        )
        if data_index is None:
            continue
        values = [value.strip() for value in lines[data_index].split("|")]
        if len(values) != len(headers):
            continue
        for position, header in enumerate(headers):
            if not header:
                continue
            value = values[position]
            if not value:
                continue
            excerpt = f"{header}: {value}"
            field = header_map.get(position)
            if field:
                _set_field(
                    fields,
                    field,
                    _entry(
                        field,
                        value,
                        0.94,
                        document_kind="pdf",
                        locator=f"page {page_number}, table line {data_index + 1}",
                        excerpt=excerpt,
                        method="delimited_table",
                    ),
                )
            else:
                _set_informational(
                    informational,
                    header,
                    _informational_entry(
                        value,
                        0.9,
                        document_kind="pdf",
                        locator=f"page {page_number}, table line {data_index + 1}",
                        excerpt=excerpt,
                        method="delimited_table",
                    ),
                )


def _pdf_quote(
    filename: str, data: bytes, *, time_budget_s: float | None = None
) -> dict[str, Any]:
    started = time.perf_counter()
    deadline = _deadline(time_budget_s)
    try:
        reader = PdfReader(io.BytesIO(data), strict=True)
    except Exception as exc:  # noqa: BLE001
        raise QuoteParseError("无法读取 PDF 报价") from exc
    if reader.is_encrypted:
        raise QuoteParseError("不支持加密 PDF 报价")
    if len(reader.pages) > MAX_PDF_PAGES:
        raise QuoteParseError(f"PDF 不得超过 {MAX_PDF_PAGES} 页")
    fields: dict[str, Any] = {}
    informational: dict[str, Any] = {}
    all_text: list[str] = []
    total_chars = 0
    for page_number, page in enumerate(reader.pages, start=1):
        _check_deadline(deadline, what="PDF")
        try:
            text = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            raise QuoteParseError(f"PDF 第 {page_number} 页文本提取失败") from exc
        total_chars += len(text)
        if total_chars > MAX_EXTRACTED_CHARS:
            raise QuoteParseError("PDF 提取文本过大")
        all_text.append(text)
        lines = text.splitlines()
        _pdf_delimited_tables(fields, informational, lines, page_number)
        for line_number, line in enumerate(lines, start=1):
            if line_number % 200 == 0:
                _check_deadline(deadline, what="PDF")
            clean = line.strip()
            if not clean:
                continue
            match = re.match(r"^\s*([^:：]{1,40})\s*[:：]\s*(.+?)\s*$", clean)
            if not match:
                continue
            label, value = match.groups()
            field = ALIASES.get(_key(label))
            if field:
                _set_field(
                    fields,
                    field,
                    _entry(
                        field,
                        value,
                        0.94,
                        document_kind="pdf",
                        locator=f"page {page_number}, line {line_number}",
                        excerpt=clean,
                        method="labelled_text",
                    ),
                )
            else:
                _set_informational(
                    informational,
                    label,
                    _informational_entry(
                        value,
                        0.9,
                        document_kind="pdf",
                        locator=f"page {page_number}, line {line_number}",
                        excerpt=clean,
                        method="labelled_text",
                    ),
                )
    full_text = "\n".join(all_text)
    if not full_text.strip():
        raise QuoteParseError("PDF 不含可提取文本；扫描件 OCR 不在当前范围内")
    _infer_common(fields, full_text, "pdf")
    return _finalize(
        fields,
        filename=filename,
        document_kind="pdf",
        processing_ms=(time.perf_counter() - started) * 1000,
        informational=informational,
    )


def parse_quote(
    filename: str, data: bytes, *, time_budget_s: float | None = None
) -> dict[str, Any]:
    suffix = _validate_file(filename, data)
    if suffix == ".xlsx":
        return _xlsx_quote(filename, data, time_budget_s=time_budget_s)
    return _pdf_quote(filename, data, time_budget_s=time_budget_s)


def fields_requiring_review(extracted: dict[str, Any]) -> list[str]:
    fields = extracted.get("fields") if isinstance(extracted, dict) else None
    if not isinstance(fields, dict):
        return list(FIELD_META)
    required = [field for field, meta in FIELD_META.items() if meta["required"]]
    if fields.get("shipping_included", {}).get("value") is False:
        required.append("shipping_fee")
    required.extend(
        field
        for field, entry in fields.items()
        if isinstance(entry, dict) and entry.get("status") == "needs_review"
    )
    return [
        field
        for field in dict.fromkeys(required)
        if fields.get(field, {}).get("value") is None
        or fields.get(field, {}).get("status") == "needs_review"
        or float(fields.get(field, {}).get("confidence", 0)) < REVIEW_THRESHOLD
    ]


__all__ = [
    "FIELD_META",
    "MAX_FILE_BYTES",
    "PARSER_VERSION",
    "QuoteParseError",
    "coerce_field_value",
    "fields_requiring_review",
    "parse_quote",
]
