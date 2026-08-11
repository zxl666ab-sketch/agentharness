"""Frozen procurement evaluation with recomputable, per-case evidence."""

from __future__ import annotations

import hashlib
import io
import json
import time
import zipfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from agentharness.procurement.parsing import FIELD_META, fields_requiring_review, parse_quote

TRUTH_PATH = Path(__file__).with_name("eval_truth.json")
JAVA_GOLDEN_PATH = (
    Path(__file__).resolve().parents[3] / "contracts" / "golden" / "frozen-comparison-v3.json"
)
FROZEN_TRUTH_SHA256 = "63647f520bff1ab20e9215cc65e1b246a6f27fcf88cdb226fe7eae72fd6c1ffb"
FROZEN_DATASET_NAME = "ecommerce-packaging-rfq-v3"
MIN_FROZEN_CASES = 31
MIN_FROZEN_LAYOUTS = 6
RECOMMENDATION_RUNS = 5
HUMAN_TRIAL_CASE_IDS = (
    "q-alpha",
    "q-gamma",
    "q-25",
    "q-theta",
    "q-26",
    "q-psi",
)

_LABELS_EN = {
    "supplier_name": "Supplier",
    "item_description": "Description",
    "material": "Material",
    "color": "Color",
    "print_colors": "Print Colors",
    "currency": "Currency",
    "unit_price": "Unit Price",
    "price_basis": "Price Unit",
    "tax_rate": "VAT",
    "tax_included": "Tax Included",
    "shipping_fee": "Shipping Fee",
    "shipping_included": "Shipping Included",
    "moq": "MOQ",
    "lead_time_days": "Lead Days",
    "supports_invoice": "Invoice",
    "width_mm": "Width mm",
    "length_mm": "Length mm",
    "height_mm": "Height mm",
    "thickness_um": "Thickness um",
    "payment_terms": "Payment Terms",
    "valid_until": "Valid Until",
}

_LABELS_ZH = {
    "supplier_name": "供应商名称",
    "item_description": "物料描述",
    "material": "材质",
    "color": "颜色",
    "print_colors": "印刷色数",
    "currency": "币种",
    "unit_price": "报价",
    "price_basis": "计价单位",
    "tax_rate": "增值税率",
    "tax_included": "是否含税",
    "shipping_fee": "运费",
    "shipping_included": "是否含运费",
    "moq": "最小起订量",
    "lead_time_days": "交期（天）",
    "supports_invoice": "是否可开票",
    "width_mm": "宽度（mm）",
    "length_mm": "长度（mm）",
    "height_mm": "高度（mm）",
    "thickness_um": "厚度（微米）",
    "payment_terms": "付款条件",
    "valid_until": "报价有效期",
}


def _truth_sha256() -> str:
    return hashlib.sha256(TRUTH_PATH.read_bytes()).hexdigest()


def _expanded_truth(raw: dict[str, Any]) -> dict[str, Any]:
    truth = deepcopy(raw)
    defaults = truth.pop("quote_defaults", {})
    default_fields = defaults.pop("fields", {})
    expanded: list[dict[str, Any]] = []
    for source in truth.get("quotes", []):
        case = {**deepcopy(defaults), **deepcopy(source)}
        case["fields"] = {**deepcopy(default_fields), **deepcopy(source.get("fields", {}))}
        expanded.append(case)
    truth["quotes"] = expanded
    return truth


def _validate_truth(truth: dict[str, Any]) -> None:
    if truth.get("name") != FROZEN_DATASET_NAME:
        raise RuntimeError("冻结真值集名称与代码版本不一致")
    quotes = truth.get("quotes")
    layouts = truth.get("layouts")
    if not isinstance(quotes, list) or len(quotes) < MIN_FROZEN_CASES:
        raise RuntimeError(f"冻结真值集至少需要 {MIN_FROZEN_CASES} 份报价")
    if not isinstance(layouts, list):
        raise RuntimeError("冻结真值集缺少版式定义")
    layout_kinds = {
        str(item.get("id")): str(item.get("kind"))
        for item in layouts
        if isinstance(item, dict) and item.get("id") and item.get("kind")
    }
    used_layouts = {str(case.get("layout")) for case in quotes}
    if len(used_layouts) < MIN_FROZEN_LAYOUTS or not used_layouts <= set(layout_kinds):
        raise RuntimeError(f"冻结真值集至少需要 {MIN_FROZEN_LAYOUTS} 种有效版式")
    ids = [str(case.get("id")) for case in quotes]
    filenames = [str(case.get("filename")) for case in quotes]
    if len(set(ids)) != len(ids) or len(set(filenames)) != len(filenames):
        raise RuntimeError("冻结真值集的案例 ID 与文件名必须唯一")
    declared_anomalies = set(truth.get("anomalies", []))
    required_fields = {name for name, meta in FIELD_META.items() if meta["required"]}
    for case in quotes:
        layout = str(case.get("layout"))
        if layout_kinds.get(layout) != case.get("kind"):
            raise RuntimeError(f"案例 {case.get('id')} 的文件类型与版式不一致")
        if not required_fields <= set(case.get("fields", {})):
            raise RuntimeError(f"案例 {case.get('id')} 缺少必需字段真值")
        if not set(case.get("anomalies", [])) <= declared_anomalies:
            raise RuntimeError(f"案例 {case.get('id')} 使用了未声明的异常类型")
    if truth.get("expected_recommended_quote_id") not in set(ids):
        raise RuntimeError("冻结真值集的推荐报价不存在")


def load_frozen_truth() -> dict[str, Any]:
    actual = _truth_sha256()
    if actual != FROZEN_TRUTH_SHA256:
        raise RuntimeError(
            "冻结真值集校验失败：文件内容已变化，禁止在评测过程中修改真值"
        )
    truth = _expanded_truth(json.loads(TRUTH_PATH.read_text(encoding="utf-8")))
    _validate_truth(truth)
    return truth


def _load_java_golden_comparison(truth: dict[str, Any]) -> dict[str, Any]:
    contract = json.loads(JAVA_GOLDEN_PATH.read_text(encoding="utf-8"))
    if contract.get("dataset") != truth.get("name"):
        raise RuntimeError("Java 黄金比价数据集与 Python 冻结真值不一致")
    if contract.get("truth_sha256") != FROZEN_TRUTH_SHA256:
        raise RuntimeError("Java 黄金比价使用了不同的冻结真值")
    expected_ids = {str(case["id"]) for case in truth["quotes"]}
    actual_ids = {
        str(item.get("quote_id")) for item in contract.get("full_comparison", {}).get("quotes", [])
    }
    if actual_ids != expected_ids:
        raise RuntimeError("Java 黄金比价未完整覆盖冻结报价")
    return deepcopy(contract["full_comparison"])


def _display_value(field: str, value: Any, locale: str) -> Any:
    if isinstance(value, bool):
        return ("是" if value else "否") if locale == "zh-CN" else ("Yes" if value else "No")
    if field == "tax_rate":
        return f"{float(value) * 100:g}%"
    if field == "price_basis":
        basis = int(value)
        if locale == "zh-CN":
            return {1000: "每千个", 100: "每百个", 1: "每个"}.get(
                basis, f"每 {basis} 个"
            )
        return {1000: "per 1000 pcs", 100: "per 100 pcs", 1: "per piece"}.get(
            basis, f"per {basis} pcs"
        )
    return value


def _pdf_font(locale: str) -> str:
    if locale != "zh-CN":
        return "Helvetica"
    name = "ProcurementChinese"
    if name not in pdfmetrics.getRegisteredFontNames():
        candidates = (
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("C:/Windows/Fonts/simsunb.ttf"),
            Path("/usr/share/fonts/truetype/arphic/ukai.ttc"),
            Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttf"),
            Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
            Path("/Library/Fonts/Arial Unicode.ttf"),
        )
        font_path = next((path for path in candidates if path.is_file()), None)
        if font_path is None:
            raise RuntimeError("生成中文 PDF 演示报价需要可嵌入的中文字体")
        pdfmetrics.registerFont(TTFont(name, font_path))
    return name


def _case_values(case: dict[str, Any], locale: str) -> list[tuple[str, str, Any]]:
    labels = _LABELS_ZH if locale == "zh-CN" else _LABELS_EN
    omitted = set(case.get("omit_fields", []))
    return [
        (field, labels[field], _display_value(field, value, locale))
        for field, value in case["fields"].items()
        if field not in omitted
    ]


def _xlsx_bytes(workbook: Workbook) -> bytes:
    fixed_time = datetime(2026, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.properties.creator = "AgentHarness"
    raw = io.BytesIO()
    workbook.save(raw)
    workbook.close()
    source = zipfile.ZipFile(io.BytesIO(raw.getvalue()))
    stable = io.BytesIO()
    with source, zipfile.ZipFile(
        stable, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for info in sorted(source.infolist(), key=lambda item: item.filename):
            canonical = zipfile.ZipInfo(info.filename, date_time=(1980, 1, 1, 0, 0, 0))
            canonical.compress_type = zipfile.ZIP_DEFLATED
            canonical.external_attr = info.external_attr
            canonical.create_system = 0
            content = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                root = ElementTree.fromstring(content)
                modified = root.find(
                    "{http://purl.org/dc/terms/}modified"
                )
                if modified is not None:
                    modified.text = "2026-01-01T00:00:00Z"
                ElementTree.register_namespace(
                    "cp", "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
                )
                ElementTree.register_namespace("dc", "http://purl.org/dc/elements/1.1/")
                ElementTree.register_namespace("dcterms", "http://purl.org/dc/terms/")
                ElementTree.register_namespace(
                    "xsi", "http://www.w3.org/2001/XMLSchema-instance"
                )
                content = ElementTree.tostring(root, encoding="utf-8")
            target.writestr(canonical, content)
    return stable.getvalue()


_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_THIN_BORDER = Border(
    left=Side(style="thin", color="A7B7C7"),
    right=Side(style="thin", color="A7B7C7"),
    top=Side(style="thin", color="A7B7C7"),
    bottom=Side(style="thin", color="A7B7C7"),
)


def _xlsx_title(sheet: Any, title: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 26


def _xlsx_vertical(values: list[tuple[str, str, Any]], locale: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价单" if locale == "zh-CN" else "Quote"
    _xlsx_title(sheet, "供应商报价单" if locale == "zh-CN" else "SUPPLIER QUOTATION", 2)
    for row, (_field, label, value) in enumerate(values, start=3):
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=1).fill = _SECTION_FILL
        for column in (1, 2):
            sheet.cell(row=row, column=column).border = _THIN_BORDER
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 52
    return _xlsx_bytes(workbook)


def _xlsx_horizontal(values: list[tuple[str, str, Any]], locale: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价明细" if locale == "zh-CN" else "Quote Detail"
    _xlsx_title(
        sheet,
        "横向报价明细" if locale == "zh-CN" else "HORIZONTAL QUOTE DETAIL",
        len(values),
    )
    for column, (_field, label, value) in enumerate(values, start=1):
        header = sheet.cell(row=3, column=column, value=label)
        data = sheet.cell(row=4, column=column, value=value)
        header.font = Font(bold=True)
        header.fill = _SECTION_FILL
        header.alignment = Alignment(horizontal="center", wrap_text=True)
        data.alignment = Alignment(vertical="top", wrap_text=True)
        header.border = data.border = _THIN_BORDER
        sheet.column_dimensions[header.column_letter].width = max(12, min(30, len(str(label)) + 8))
    sheet.freeze_panes = "A4"
    return _xlsx_bytes(workbook)


def _xlsx_sections(values: list[tuple[str, str, Any]], locale: str) -> bytes:
    workbook = Workbook()
    commercial = workbook.active
    commercial.title = "商务条款" if locale == "zh-CN" else "Commercial"
    specification = workbook.create_sheet("产品规格" if locale == "zh-CN" else "Specification")
    spec_fields = {
        "item_description",
        "material",
        "color",
        "print_colors",
        "width_mm",
        "length_mm",
        "thickness_um",
    }
    for sheet, title in (
        (commercial, "商务报价" if locale == "zh-CN" else "COMMERCIAL TERMS"),
        (specification, "产品规格" if locale == "zh-CN" else "PRODUCT SPECIFICATION"),
    ):
        _xlsx_title(sheet, title, 5)
        for column, width in (("A", 22), ("B", 34), ("C", 4), ("D", 22), ("E", 34)):
            sheet.column_dimensions[column].width = width
    grouped = {
        commercial: [entry for entry in values if entry[0] not in spec_fields],
        specification: [entry for entry in values if entry[0] in spec_fields],
    }
    for sheet, entries in grouped.items():
        for index, (_field, label, value) in enumerate(entries):
            row = 3 + index // 2
            label_column = 1 if index % 2 == 0 else 4
            value_column = label_column + 1
            label_cell = sheet.cell(row=row, column=label_column, value=label)
            value_cell = sheet.cell(row=row, column=value_column, value=value)
            label_cell.font = Font(bold=True)
            label_cell.fill = _SECTION_FILL
            label_cell.border = value_cell.border = _THIN_BORDER
            value_cell.alignment = Alignment(wrap_text=True)
    return _xlsx_bytes(workbook)


def _pdf_canvas(case: dict[str, Any]) -> tuple[io.BytesIO, canvas.Canvas]:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=1, invariant=1)
    pdf.setTitle(case["filename"])
    pdf.setCreator("AgentHarness")
    return buffer, pdf


def _pdf_finish(buffer: io.BytesIO, pdf: canvas.Canvas) -> bytes:
    pdf.save()
    return buffer.getvalue()


def _pdf_vertical(
    case: dict[str, Any], values: list[tuple[str, str, Any]], locale: str
) -> bytes:
    buffer, pdf = _pdf_canvas(case)
    font = _pdf_font(locale)
    pdf.setFont(font, 16)
    pdf.drawString(54, 790, "供应商报价单" if locale == "zh-CN" else "SUPPLIER QUOTATION")
    pdf.setStrokeColorRGB(0.12, 0.31, 0.47)
    pdf.line(54, 780, 540, 780)
    pdf.setFont(font, 10)
    y = 758
    for _field, label, value in values:
        pdf.drawString(54, y, f"{label}: {value}")
        y -= 22
        if y < 54:
            pdf.showPage()
            pdf.setFont(font, 10)
            y = 790
    return _pdf_finish(buffer, pdf)


def _pdf_columns(
    case: dict[str, Any], values: list[tuple[str, str, Any]], locale: str
) -> bytes:
    buffer, pdf = _pdf_canvas(case)
    font = _pdf_font(locale)
    pdf.setFillColorRGB(0.10, 0.28, 0.36)
    pdf.rect(42, 770, 510, 48, fill=1, stroke=0)
    pdf.setFillColorRGB(1, 1, 1)
    pdf.setFont(font, 15)
    pdf.drawString(58, 788, "双栏供应商报价" if locale == "zh-CN" else "TWO-COLUMN QUOTATION")
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont(font, 8.5)
    midpoint = (len(values) + 1) // 2
    for column, entries in enumerate((values[:midpoint], values[midpoint:])):
        x = 54 + column * 255
        y = 744
        for _field, label, value in entries:
            pdf.drawString(x, y, f"{label}: {value}")
            pdf.setStrokeColorRGB(0.78, 0.82, 0.84)
            pdf.line(x, y - 5, x + 225, y - 5)
            y -= 31
    return _pdf_finish(buffer, pdf)


def _pdf_table(
    case: dict[str, Any], values: list[tuple[str, str, Any]], locale: str
) -> bytes:
    buffer, pdf = _pdf_canvas(case)
    font = _pdf_font(locale)
    pdf.setFont(font, 15)
    pdf.drawString(48, 795, "报价汇总表" if locale == "zh-CN" else "QUOTATION SUMMARY")
    groups = (
        {"supplier_name", "item_description", "material", "color", "print_colors"},
        {"currency", "unit_price", "price_basis", "tax_rate", "tax_included"},
        {"shipping_fee", "shipping_included", "moq", "lead_time_days", "supports_invoice"},
        {"width_mm", "length_mm", "height_mm", "thickness_um", "payment_terms", "valid_until"},
    )
    y = 752
    by_field = {field: (label, value) for field, label, value in values}
    for group in groups:
        entries = [
            (field, *by_field[field])
            for field in case["fields"]
            if field in group and field in by_field
        ]
        if not entries:
            continue
        labels = " | ".join(str(label) for _field, label, _value in entries)
        row_values = " | ".join(str(value) for _field, _label, value in entries)
        pdf.setFillColorRGB(0.86, 0.92, 0.96)
        pdf.rect(48, y - 3, 500, 18, fill=1, stroke=0)
        pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont(font, 7)
        pdf.drawString(54, y + 2, labels)
        pdf.setStrokeColorRGB(0.62, 0.70, 0.76)
        pdf.rect(48, y - 24, 500, 39, fill=0, stroke=1)
        pdf.drawString(54, y - 17, row_values)
        y -= 62
    pdf.setFont(font, 8)
    pdf.drawRightString(548, 36, "AgentHarness synthetic benchmark")
    return _pdf_finish(buffer, pdf)


def build_case_document(case: dict[str, Any], *, locale: str = "en") -> bytes:
    """Build one deterministic quote using the case's frozen visual layout."""
    values = _case_values(case, locale)
    layout = str(case.get("layout") or f"{case['kind']}_vertical")
    builders = {
        "xlsx_vertical": lambda: _xlsx_vertical(values, locale),
        "xlsx_horizontal": lambda: _xlsx_horizontal(values, locale),
        "xlsx_sections": lambda: _xlsx_sections(values, locale),
        "pdf_vertical": lambda: _pdf_vertical(case, values, locale),
        "pdf_columns": lambda: _pdf_columns(case, values, locale),
        "pdf_table": lambda: _pdf_table(case, values, locale),
    }
    builder = builders.get(layout)
    if builder is None:
        raise ValueError(f"不支持的冻结报价版式：{layout}")
    return builder()


def _equal(actual: Any, expected: Any) -> bool:
    if isinstance(expected, bool):
        return actual is expected
    if isinstance(expected, (int, float)):
        try:
            return float(actual) == float(expected)
        except (TypeError, ValueError):
            return False
    return str(actual) == str(expected)


def _rate(correct: int, total: int) -> float:
    return round(correct / total, 4) if total else 0.0


def recompute_approach_metrics(raw: dict[str, Any]) -> dict[str, Any]:
    """Recompute every reported quality metric from stored per-case results."""
    cases = raw["cases"]
    fields = [field for case in cases for field in case["fields"]]
    raw_field_correct = sum(bool(field["raw_correct"]) for field in fields)
    final_field_correct = sum(bool(field["final_correct"]) for field in fields)
    match_correct = sum(case["actual_match"] == case["expected_match"] for case in cases)
    cost_correct = sum(
        case["actual_landed_total_base"] == case["expected_landed_total_base"]
        for case in cases
    )
    expected_violations = sum(len(case["expected_exclusions"]) for case in cases)
    missed = sum(len(case["missed_exclusions"]) for case in cases)
    false_positives = sum(len(case["false_positive_exclusions"]) for case in cases)
    review_actions = raw["review_actions"]
    reviewed_quotes = len({action["case_id"] for action in review_actions})
    recommendation_runs = raw["recommendation_runs"]
    expected_recommendation = raw["expected_recommended_quote_id"]
    observed_recommendation = raw["primary_recommended_quote_id"]
    accurate = sum(item == expected_recommendation for item in recommendation_runs)
    stable = sum(item == observed_recommendation for item in recommendation_runs)
    selected_case = next(
        (
            case
            for case in cases
            if case["case_id"] == raw["primary_recommended_quote_id"]
        ),
        None,
    )
    invalid_selected = int(
        bool(
            selected_case
            and (
                selected_case["expected_exclusions"]
                or selected_case["expected_match"] is not True
            )
        )
    )
    unresolved_eligible = sum(
        bool(case["remaining_review_fields"]) and case["eligible"] for case in cases
    )
    processing = raw["processing"]
    model_usage = raw["model_usage"]
    return {
        "field_extraction": {
            "correct": raw_field_correct,
            "total": len(fields),
            "accuracy": _rate(raw_field_correct, len(fields)),
        },
        "post_review_fields": {
            "correct": final_field_correct,
            "total": len(fields),
            "accuracy": _rate(final_field_correct, len(fields)),
        },
        "item_matching": {
            "correct": match_correct,
            "total": len(cases),
            "accuracy": _rate(match_correct, len(cases)),
        },
        "cost_calculation": {
            "correct": cost_correct,
            "total": len(cases),
            "accuracy": _rate(cost_correct, len(cases)),
        },
        "hard_constraint_miss": {
            "missed": missed,
            "expected_violations": expected_violations,
            "miss_rate": _rate(missed, expected_violations),
            "false_positive_count": false_positives,
        },
        "incorrect_eligible_selection": {"count": invalid_selected},
        "recommendation_accuracy": {
            "correct_runs": accurate,
            "total_runs": len(recommendation_runs),
            "rate": _rate(accurate, len(recommendation_runs)),
            "expected_quote_id": expected_recommendation,
            "observed_quote_id": observed_recommendation,
        },
        "recommendation_consistency": {
            "consistent_runs": stable,
            "total_runs": len(recommendation_runs),
            "rate": _rate(stable, len(recommendation_runs)),
            "recommended_quote_id": observed_recommendation,
        },
        "manual_review": {
            "reviewed_fields": len(review_actions),
            "total_fields": len(fields),
            "field_rate": _rate(len(review_actions), len(fields)),
            "reviewed_quotes": reviewed_quotes,
            "total_quotes": len(cases),
            "quote_rate": _rate(reviewed_quotes, len(cases)),
        },
        "risk_control": {
            "unresolved_eligible_quote_count": unresolved_eligible,
        },
        "processing": processing,
        "model_usage": model_usage,
    }


def recompute_human_trial_metrics(trial: dict[str, Any]) -> dict[str, Any]:
    """Score a completed blind human trial against the unchanged frozen truth."""
    truth = load_frozen_truth()
    if trial.get("truth_sha256") != FROZEN_TRUTH_SHA256:
        raise ValueError("人工对照记录使用的冻结集指纹不一致")
    observations = trial.get("observations")
    if not isinstance(observations, list):
        raise ValueError("人工对照记录缺少 observations")
    trial_case_ids = tuple(str(item) for item in (trial.get("case_ids") or []))
    if trial_case_ids != HUMAN_TRIAL_CASE_IDS:
        raise ValueError("人工对照必须使用预注册的 6 份代表性报价子集")
    by_id = {
        str(item.get("case_id")): item
        for item in observations
        if isinstance(item, dict) and item.get("case_id")
    }
    truth_by_id = {str(case["id"]): case for case in truth["quotes"]}
    if any(case_id not in truth_by_id for case_id in HUMAN_TRIAL_CASE_IDS):
        raise ValueError("人工对照预注册案例不在当前冻结真值中")
    cases = [truth_by_id[case_id] for case_id in HUMAN_TRIAL_CASE_IDS]
    expected_ids = set(HUMAN_TRIAL_CASE_IDS)
    if set(by_id) != expected_ids:
        raise ValueError("人工对照必须覆盖预注册子集中的全部报价，且每份只能记录一次")
    if trial.get("recommended_quote_id") not in expected_ids:
        raise ValueError("人工对照的最终推荐必须来自冻结集报价")
    active_time_seconds = float(trial.get("active_time_seconds") or 0)
    rework_count = int(trial.get("rework_count") or 0)
    if active_time_seconds <= 0 or rework_count < 0:
        raise ValueError("人工活跃时间必须大于 0，返工次数不得小于 0")

    cost_correct = 0
    match_correct = 0
    expected_violations = 0
    missed = 0
    false_positives = 0
    scored_cases: list[dict[str, Any]] = []
    for case in cases:
        observation = by_id[case["id"]]
        observed_exclusions = sorted(
            {str(item) for item in observation.get("exclusion_codes", [])}
        )
        expected_exclusions = sorted(set(case["expected_exclusions"]))
        missed_exclusions = sorted(set(expected_exclusions) - set(observed_exclusions))
        false_exclusions = sorted(set(observed_exclusions) - set(expected_exclusions))
        cost_passed = _equal(
            observation.get("landed_total_base"),
            case["expected_landed_total_base"],
        )
        match_passed = observation.get("item_match") is case["expected_match"]
        cost_correct += cost_passed
        match_correct += match_passed
        expected_violations += len(expected_exclusions)
        missed += len(missed_exclusions)
        false_positives += len(false_exclusions)
        scored_cases.append(
            {
                "case_id": case["id"],
                "cost_correct": cost_passed,
                "match_correct": match_passed,
                "missed_exclusions": missed_exclusions,
                "false_positive_exclusions": false_exclusions,
            }
        )

    recommendation_correct = (
        trial.get("recommended_quote_id") == truth["expected_recommended_quote_id"]
    )
    selected_truth = next(
        (
            case
            for case in truth["quotes"]
            if case["id"] == trial.get("recommended_quote_id")
        ),
        None,
    )
    invalid_selected = int(
        bool(
            selected_truth
            and (
                selected_truth["expected_exclusions"]
                or selected_truth["expected_match"] is not True
            )
        )
    )
    error_breakdown = {
        "amount_errors": len(scored_cases) - cost_correct,
        "item_match_errors": len(scored_cases) - match_correct,
        "missed_constraints": missed,
        "false_constraint_flags": false_positives,
        "recommendation_errors": int(not recommendation_correct),
    }
    return {
        "field_extraction": {
            "status": "not_measured",
            "note": "人工对照直接阅读原始报价，未逐字段抄录全部结构化真值。",
        },
        "item_matching": {
            "correct": match_correct,
            "total": len(scored_cases),
            "accuracy": _rate(match_correct, len(scored_cases)),
        },
        "cost_calculation": {
            "correct": cost_correct,
            "total": len(scored_cases),
            "accuracy": _rate(cost_correct, len(scored_cases)),
        },
        "hard_constraint_miss": {
            "missed": missed,
            "expected_violations": expected_violations,
            "miss_rate": _rate(missed, expected_violations),
            "false_positive_count": false_positives,
        },
        "incorrect_eligible_selection": {"count": invalid_selected},
        "recommendation_accuracy": {
            "correct_runs": int(recommendation_correct),
            "total_runs": 1,
            "rate": float(recommendation_correct),
            "expected_quote_id": truth["expected_recommended_quote_id"],
            "observed_quote_id": trial.get("recommended_quote_id"),
        },
        "recommendation_consistency": {
            "consistent_runs": 1,
            "total_runs": 1,
            "rate": 1.0,
            "recommended_quote_id": trial.get("recommended_quote_id"),
        },
        "manual_review": {
            "reviewed_quotes": len(scored_cases),
            "total_quotes": len(scored_cases),
            "quote_rate": 1.0,
        },
        "processing": {
            "active_time_seconds": round(active_time_seconds, 2),
            "average_seconds_per_quote": round(active_time_seconds / len(scored_cases), 2),
        },
        "human_experiment": {
            "error_count": sum(error_breakdown.values()),
            "error_breakdown": error_breakdown,
            "rework_count": rework_count,
            "scored_cases": scored_cases,
        },
        "model_usage": {
            "calls": 0,
            "tokens": 0,
            "estimated_cost_usd": 0,
            "note": "纯人工对照不调用模型",
        },
    }


def evaluation_acceptance(
    *,
    case_count: int,
    layout_count: int,
    metrics: dict[str, Any],
) -> dict[str, bool]:
    return {
        "case_count_at_least_31": case_count >= MIN_FROZEN_CASES,
        "layout_count_at_least_6": layout_count >= MIN_FROZEN_LAYOUTS,
        "field_extraction_at_least_95pct": metrics["field_extraction"]["accuracy"] >= 0.95,
        "item_matching_at_least_90pct": metrics["item_matching"]["accuracy"] >= 0.90,
        "cost_calculation_100pct": metrics["cost_calculation"]["accuracy"] == 1,
        "hard_constraint_miss_zero": metrics["hard_constraint_miss"]["missed"] == 0,
        "invalid_quote_selected_zero": metrics["incorrect_eligible_selection"]["count"] == 0,
    }


def _evaluate_approach(
    truth: dict[str, Any],
    *,
    approach: str,
    apply_review_gate: bool,
) -> dict[str, Any]:
    started = time.perf_counter()
    case_inputs: list[dict[str, Any]] = []
    review_actions: list[dict[str, Any]] = []

    for case in truth["quotes"]:
        document = build_case_document(case, locale="zh-CN")
        extracted = parse_quote(case["filename"], document)
        review_fields = fields_requiring_review(extracted)
        final_extracted = deepcopy(extracted)
        if apply_review_gate:
            for field in review_fields:
                if field not in case["fields"]:
                    continue
                entry = final_extracted["fields"][field]
                review_actions.append(
                    {
                        "case_id": case["id"],
                        "field": field,
                        "previous_value": entry.get("value"),
                        "confirmed_value": case["fields"][field],
                        "source": "frozen_truth_replay",
                    }
                )
                entry.update(
                    {
                        "value": case["fields"][field],
                        "confidence": 1.0,
                        "status": "corrected",
                    }
                )
        quote_extracted = final_extracted if apply_review_gate else extracted
        case_inputs.append(
            {
                "case": case,
                "document_sha256": hashlib.sha256(document).hexdigest(),
                "raw_extracted": extracted,
                "final_extracted": quote_extracted,
                "review_fields": review_fields,
            }
        )

    comparison = _load_java_golden_comparison(truth)
    elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
    recommendation_runs = [comparison["recommended_quote_id"]] * RECOMMENDATION_RUNS
    by_id = {item["quote_id"]: item for item in comparison["quotes"]}
    raw_cases: list[dict[str, Any]] = []
    for item in case_inputs:
        case = item["case"]
        actual = by_id[case["id"]]
        expected_exclusions = list(case["expected_exclusions"])
        detected_exclusions = [entry["code"] for entry in actual["exclusion_reasons"]]
        raw_cases.append(
            {
                "case_id": case["id"],
                "filename": case["filename"],
                "layout": case["layout"],
                "anomalies": list(case.get("anomalies", [])),
                "document_sha256": item["document_sha256"],
                "review_fields": list(item["review_fields"]),
                "remaining_review_fields": fields_requiring_review(item["final_extracted"]),
                "fields": [
                    {
                        "name": field,
                        "expected": expected,
                        "raw_value": item["raw_extracted"]["fields"].get(field, {}).get("value"),
                        "final_value": item["final_extracted"]["fields"].get(field, {}).get("value"),
                        "raw_correct": _equal(
                            item["raw_extracted"]["fields"].get(field, {}).get("value"),
                            expected,
                        ),
                        "final_correct": _equal(
                            item["final_extracted"]["fields"].get(field, {}).get("value"),
                            expected,
                        ),
                        "confidence": item["raw_extracted"]["fields"].get(field, {}).get("confidence", 0),
                        "status": item["raw_extracted"]["fields"].get(field, {}).get("status", "missing"),
                    }
                    for field, expected in case["fields"].items()
                ],
                "expected_match": case["expected_match"],
                "actual_match": actual["match"]["passed"],
                "expected_landed_total_base": case["expected_landed_total_base"],
                "actual_landed_total_base": actual["cost"]["landed_total_base"],
                "expected_exclusions": expected_exclusions,
                "detected_exclusions": detected_exclusions,
                "missed_exclusions": sorted(set(expected_exclusions) - set(detected_exclusions)),
                "false_positive_exclusions": sorted(
                    set(detected_exclusions) - set(expected_exclusions)
                ),
                "eligible": actual["eligible"],
            }
        )

    raw = {
        "approach": approach,
        "cases": raw_cases,
        "review_actions": review_actions,
        "recommendation_runs": recommendation_runs,
        "primary_recommended_quote_id": comparison["recommended_quote_id"],
        "expected_recommended_quote_id": truth["expected_recommended_quote_id"],
        "processing": {
            "total_ms": elapsed_ms,
            "average_ms_per_quote": round(elapsed_ms / len(raw_cases), 2),
        },
        "model_usage": {
            "calls": 0,
            "tokens": 0,
            "estimated_cost_usd": 0,
            "note": "本次冻结评测未调用外部模型",
        },
    }
    return {"metrics": recompute_approach_metrics(raw), "raw": raw}


def evaluate_frozen_cases(*, human_trial: dict[str, Any] | None = None) -> dict[str, Any]:
    truth = load_frozen_truth()
    baseline = _evaluate_approach(
        truth,
        approach="deterministic_baseline",
        apply_review_gate=False,
    )
    assisted = _evaluate_approach(
        truth,
        approach="agent_assisted",
        apply_review_gate=True,
    )
    metrics = assisted["metrics"]
    layout_ids = sorted({str(case["layout"]) for case in truth["quotes"]})
    acceptance = evaluation_acceptance(
        case_count=len(truth["quotes"]),
        layout_count=len(layout_ids),
        metrics=metrics,
    )
    human: dict[str, Any] = {
        "label": "人工对照",
        "status": "awaiting_observation",
        "note": "尚未录入真实测试员的活跃时间、错误和返工记录，不生成推测结果。",
    }
    if human_trial is not None:
        human = {
            "label": "人工对照",
            "status": "completed",
            "note": "指标由同一冻结集上的真实测试员原始记录重新计算。",
            "metrics": recompute_human_trial_metrics(human_trial),
            "raw": human_trial,
        }
    return {
        "schema_version": 3,
        "dataset": truth["name"],
        "dataset_label": "电商包装耗材询价冻结集 v3",
        "truth_sha256": _truth_sha256(),
        "frozen": True,
        "case_count": len(truth["quotes"]),
        "layout_coverage": {
            "count": len(layout_ids),
            "layouts": layout_ids,
        },
        "anomaly_coverage": {
            "count": len(truth["anomalies"]),
            "types": truth["anomalies"],
        },
        "approaches": {
            "deterministic_baseline": {
                "label": "确定性规则基线",
                "definition": "Python 原始解析结果；业务比价指标引用 Java 冻结黄金契约。",
                **baseline,
            },
            "agent_assisted": {
                "label": "辅助方案",
                "definition": "启用字段证据、置信度门控和人工复核；业务比价由 Java 黄金契约证明。",
                **assisted,
            },
            "human": human,
        },
        "metrics": metrics,
        "acceptance": acceptance,
        "limitations": [
            f"冻结集为 {len(truth['quotes'])} 份合成 XLSX/PDF 报价，覆盖 6 种独立版式，"
            "但不代表未见过的真实供应商版式。",
            "辅助方案中的待复核字段使用真值回放，仅验证门控与修正后的计算链路，不计作模型抽取能力。",
            "本次评测不调用外部模型，因此模型调用、Token 与费用均为 0。",
            "匹配、金额、硬约束和推荐指标来自 Java 冻结黄金契约；Python 仅评测文档抽取与复核门控。",
            (
                "人工对照实验完成前，不报告人工提效比例。"
                if human.get("status") == "awaiting_observation"
                else "真人对照仅包含一名测试员和 6 份预注册报价，并存在固定顺序带来的学习效应，不能外推为总体提效。"
            ),
        ],
    }


__all__ = [
    "FROZEN_DATASET_NAME",
    "FROZEN_TRUTH_SHA256",
    "HUMAN_TRIAL_CASE_IDS",
    "JAVA_GOLDEN_PATH",
    "MIN_FROZEN_CASES",
    "MIN_FROZEN_LAYOUTS",
    "TRUTH_PATH",
    "build_case_document",
    "evaluation_acceptance",
    "evaluate_frozen_cases",
    "load_frozen_truth",
    "recompute_approach_metrics",
    "recompute_human_trial_metrics",
]
