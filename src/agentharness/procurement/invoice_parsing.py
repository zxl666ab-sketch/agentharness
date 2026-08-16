"""Invoice field extraction (P3-1): deterministic, resource-bounded.

支持 XLSX 与文本型 PDF 的常见发票版式：
- 键值对（发票代码/号码、开票日期、价税合计、税额、税率、销售方等）；
- 明细行表格（数量/单价/金额）求和得到数量与不含税金额。

输出字段全部为字符串/Decimal 的规范化值，Java 侧三单匹配只认这里的结构化结果。
"""

from __future__ import annotations

import io
import re
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

INVOICE_PARSER_VERSION = "invoice-v1"
MAX_INVOICE_BYTES = 5 * 1024 * 1024
MAX_XLSX_ROWS = 500
MAX_XLSX_COLUMNS = 40
MAX_PDF_PAGES = 20
MAX_EXTRACTED_CHARS = 200_000


class InvoiceParseError(ValueError):
    """An invoice cannot be parsed within the supported boundary."""


_ALIASES: dict[str, list[str]] = {
    "invoice_code": ["发票代码", "发票代码", "invoice code", "invoicecode"],
    "invoice_no": ["发票号码", "发票号", "发票编号", "invoice no", "invoice number", "invoiceno", "invoicenumber"],
    "issue_date": ["开票日期", "发票日期", "开具日期", "issue date", "invoice date", "issuedate", "date"],
    "supplier_name": ["销售方名称", "销售方", "开票方", "开票方名称", "供应商名称", "供应商", "seller name", "seller", "supplier"],
    "quantity": ["数量", "件数", "数量合计", "quantity", "qty"],
    "unit": ["单位", "unit"],
    "unit_price": ["单价", "不含税单价", "unit price", "unitprice"],
    "amount_excluding_tax": ["金额", "不含税金额", "合计金额", "金额合计", "excl tax amount", "excltax amount", "excltax", "amount excl tax", "amount"],
    "tax_amount": ["税额", "税额合计", "增值税额", "tax amount", "taxamount"],
    "total_amount": ["价税合计", "价税总计", "合计", "总金额", "价税合计（小写）", "grand total", "total amount", "total"],
    "tax_rate": ["税率", "增值税率", "税率%", "tax rate", "vat rate", "taxrate"],
}

_AMOUNT_KEYS = {
    "amount_excluding_tax",
    "tax_amount",
    "total_amount",
    "unit_price",
    "quantity",
}


def _key(value: Any) -> str:
    return re.sub(r"[\s:：*（）()【】\[\]_\-—]", "", str(value or "")).strip().lower()


def _plain_decimal(value: Decimal) -> str:
    if value == 0:
        return "0"
    return format(value.normalize(), "f")


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    text = re.sub(r"[,\s]", "", text)
    text = re.sub(r"[¥￥]", "", text)
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def _date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    match = re.search(r"(\d{4})[年/\-.](\d{1,2})[月/\-.](\d{1,2})", text)
    if match:
        year, month, day = (int(part) for part in match.groups())
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def _entries_from_cells(rows: list[list[Any]], kind: str) -> list[dict[str, Any]]:
    """把键值单元格对转换为 {field, value, excerpt, confidence}。"""
    entries: list[dict[str, Any]] = []
    for row in rows:
        cells = [cell for cell in row if cell is not None and str(cell).strip()]
        if len(cells) < 2:
            continue
        label = _key(cells[0])
        if not label:
            continue
        field = next(
            (name for name, aliases in _ALIASES.items()
             if any(label == _key(alias) or label.startswith(_key(alias)) for alias in aliases)),
            None,
        )
        if field is None:
            continue
        value = cells[1]
        excerpt = f"{cells[0]}：{value}"
        entries.append({"field": field, "value": value, "excerpt": excerpt, "confidence": 0.95, "kind": kind})
    return entries


def _line_items(rows: list[list[Any]], kind: str) -> dict[str, Any]:
    """明细行聚合：数量合计 + 金额合计（取首个匹配列的数值行）。"""
    header_idx: dict[str, int] = {}
    totals = {"quantity": Decimal(0), "amount_excluding_tax": Decimal(0), "tax_amount": Decimal(0)}
    saw_header = False
    for row in rows:
        cells = [cell if cell is not None else "" for cell in row]
        if not saw_header:
            for index, cell in enumerate(cells):
                label = _key(cell)
                if not label:
                    continue
                field = next(
                    (name for name, aliases in _ALIASES.items()
                     if any(_key(alias) == label for alias in aliases)),
                    None,
                )
                if field in {"quantity", "amount_excluding_tax", "tax_amount", "unit_price"}:
                    header_idx[field] = index
            if len(header_idx) >= 2:
                saw_header = True
            continue
        values = {field: _decimal(cells[index]) for field, index in header_idx.items()}
        if not any(values.values()):
            continue
        if values.get("quantity"):
            totals["quantity"] += values["quantity"]
        if values.get("amount_excluding_tax"):
            totals["amount_excluding_tax"] += values["amount_excluding_tax"]
        if values.get("tax_amount"):
            totals["tax_amount"] += values["tax_amount"]
    result: dict[str, Any] = {}
    for field, total in totals.items():
        if total and total != 0:
            result[field] = {
                "value": _plain_decimal(total),
                "confidence": 0.9,
                "excerpt": f"{field}（明细行合计）",
                "kind": kind,
            }
    return result


def _from_rows(rows: list[list[Any]], kind: str) -> dict[str, Any]:
    entries = _entries_from_cells(rows, kind)
    entries.extend(_line_items(rows, kind).values())
    fields: dict[str, Any] = {}
    for entry in entries:
        field = entry["field"]
        current = fields.get(field)
        if current is None:
            fields[field] = {key: entry[key] for key in ("value", "confidence", "excerpt", "kind")}
            continue
        # 键值对优先于行合计（合计行常把"合计"一词命中 multiple aliases）
        if entry["kind"] == "key_value" or current["kind"] != "key_value":
            fields[field] = {key: entry[key] for key in ("value", "confidence", "excerpt", "kind")}
    return fields


def _xlsx_rows(data: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[list[Any]] = []
    try:
        for sheet in workbook.worksheets[:2]:
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= MAX_XLSX_ROWS:
                    break
                rows.append(list(row))
    finally:
        workbook.close()
    return rows


def _pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for page in reader.pages[:MAX_PDF_PAGES]:
        parts.append(str(page.extract_text() or ""))
    return "".join(parts)[:MAX_EXTRACTED_CHARS]


def _from_text(text: str, kind: str) -> dict[str, Any]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    fields: dict[str, Any] = {}
    for line in lines:
        matches: list[tuple[int, str, str, str, int, int]] = []  # (alias_len, alias, field, value, start, end)
        for field, aliases in _ALIASES.items():
            for alias in aliases:
                escaped = re.escape(alias).replace("\\ ", "\\s*")
                pattern = re.compile(
                    rf"{escaped}[\s:：]*([0-9,¥￥.\-年月日%]+|[^\s，。；;]{{2,40}})",
                    re.IGNORECASE,
                )
                for match in pattern.finditer(line):
                    matches.append((
                        len(alias), alias, field,
                        match.group(1).strip().rstrip("元"),
                        match.start(), match.end(),
                    ))
        matches.sort(key=lambda item: (-item[0], item[1]))
        taken: list[tuple[int, int]] = []
        for _alias_len, _alias, field, value, start, end in matches:
            if any(start < other_end and other_start < end for other_start, other_end in taken):
                continue
            taken.append((start, end))
            if field in fields:
                continue
            fields[field] = {
                "value": value,
                "confidence": 0.88,
                "excerpt": line,
                "kind": "key_value",
            }
    return fields


def parse_invoice(filename: str, data: bytes) -> dict[str, Any]:
    """解析发票，返回结构化字段（Java 三单匹配只消费 invoice 键）。"""
    if len(data) > MAX_INVOICE_BYTES:
        raise InvoiceParseError("invoice file exceeds size limit")
    started = time.monotonic()
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        rows = _xlsx_rows(data)
        fields = _from_rows(rows, "key_value")
        kind = "xlsx"
    elif lower.endswith(".pdf"):
        fields = _from_text(_pdf_text(data), "key_value")
        kind = "pdf"
    else:
        raise InvoiceParseError("invoice must be xlsx or text pdf")

    invoice: dict[str, Any] = {}
    for field, spec in {
        "invoice_code": ("text", None),
        "invoice_no": ("text", None),
        "issue_date": ("date", None),
        "supplier_name": ("text", None),
        "quantity": ("decimal", None),
        "unit": ("text", None),
        "unit_price": ("decimal", None),
        "amount_excluding_tax": ("decimal", None),
        "tax_amount": ("decimal", None),
        "total_amount": ("decimal", None),
        "tax_rate": ("rate", None),
    }.items():
        entry = fields.get(field)
        if entry is None:
            continue
        value = entry["value"]
        if spec[0] == "decimal":
            converted = _decimal(value)
            invoice[field] = None if converted is None else _plain_decimal(converted)
        elif spec[0] == "date":
            invoice[field] = _date(value)
        elif spec[0] == "rate":
            converted = _decimal(value)
            if converted is not None:
                rate = converted / Decimal("100") if converted > Decimal("1") else converted
                invoice[field] = _plain_decimal(rate)
        else:
            invoice[field] = str(value).strip()

    return {
        "schema_version": 1,
        "parser_version": INVOICE_PARSER_VERSION,
        "document_kind": kind,
        "invoice": invoice,
        "processing_ms": str(round((time.monotonic() - started) * 1000, 2)),
    }


FIELD_LABELS = {
    "quantity": "数量",
    "unit_price": "单价",
    "total_amount": "价税合计",
    "tax_rate": "税率",
}


def build_diff_explanation(diffs: list[dict[str, Any]]) -> dict[str, Any]:
    """P3-1 模式 C：Java 结构化差异 → 自然语言原因与处理建议。

    确定性模板；解释中出现的每个数字都直接取自注入的 diffs
    （满足「差异解释数值引用一致性」评测硬校验）。
    """
    if not diffs:
        raise ValueError("diff explanation requires at least one diff")
    reasons: list[str] = []
    suggestions: list[str] = []
    for diff in diffs:
        field = str(diff.get("field") or "")
        label = FIELD_LABELS.get(field, field)
        expected = str(diff.get("expected") or "")
        actual = str(diff.get("actual") or "")
        delta = str(diff.get("diff") or "")
        reasons.append(f"{label}不一致：订单/收货为 {expected}，发票为 {actual}（差异 {delta}）")
        if field == "quantity":
            suggestions.append("请核对收货数量与发票数量；如为补开票请说明原因")
        elif field == "unit_price":
            suggestions.append("请核对单价口径（含税/不含税）与订单到货单价")
        elif field == "tax_rate":
            suggestions.append("请核对税率档位与批准报价税率")
        else:
            suggestions.append("请核对金额合计与订单到货总价")
    return {
        "reason": f"三单匹配存在 {len(diffs)} 项差异：" + "；".join(reasons) + "。",
        "suggestions": suggestions,
        "source": "deterministic_agent",
    }


__all__ = [
    "FIELD_LABELS",
    "INVOICE_PARSER_VERSION",
    "InvoiceParseError",
    "build_diff_explanation",
    "parse_invoice",
]
